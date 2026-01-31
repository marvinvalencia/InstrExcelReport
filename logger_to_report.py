# logger_to_report.py
"""Convert a test-data-logger CSV (Data Instr INSTR export) into a ready-to-use Excel report.

The input CSVs from the logger are typically UTF-16, tab-delimited, and contain:
  * A metadata section (Name/Owner/Acquisition/etc.)
  * A channel definition table (one row per channel)
  * A data table where each channel has a value column and an alarm column

This script:
  * Parses metadata + channel list
  * Extracts channel readings (ignoring alarm columns)
  * Down-samples 10-second data to full elapsed minutes (keeps 0,1,2,...) 
  * Builds an .xlsx report with:
      - Absolute readings
      - Temperature rise (delta from ambient) per thermocouple
      - Mean/max/core summary series
      - Charts on a "Summary of Results" tab

Usage:
  python logger_to_report.py "input.csv" "output.xlsx"
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ParsedLoggerFile:
    metadata: Dict[str, str]
    channels: List[int]
    rows: List[Tuple[int, dt.date, dt.time, float, Dict[int, float]]]
    # rows: (scan, date, time, elapsed_minutes, {channel: value})


def _read_lines_utf16(path: Path) -> List[str]:
    """Read the logger export file.

    The INSTR export is typically UTF-16. Some files are tab-delimited, others are comma-delimited.
    We keep raw lines and detect delimiter later.
    """
    with path.open("r", encoding="utf-16") as f:
        return f.read().splitlines()


def _detect_delimiter_from_channel_header(lines: List[str]) -> str:
    """Return '\t' or ',' depending on the channel table header format."""
    for line in lines[:200]:
        s = line.strip()
        if s.startswith("Channel") and "Name" in s and "Function" in s:
            if "\t" in s and "Channel\tName\tFunction" in s:
                return "\t"
            if "," in s and "Channel,Name,Function" in s:
                return ","
    # Fallback: choose the most common delimiter among early lines
    tab_hits = sum(1 for l in lines[:50] if "\t" in l)
    comma_hits = sum(1 for l in lines[:50] if "," in l)
    return "\t" if tab_hits >= comma_hits else ","


def _split_fields(line: str, delim: str) -> List[str]:
    # Some rows have extra spaces before Control:, so trim first
    if delim == ",":
        return [p.strip() for p in line.split(",")]
    return [p.strip() for p in line.split("\t")]


def _parse_metadata(lines: List[str]) -> Dict[str, str]:
    meta: Dict[str, str] = {}
    for line in lines[:60]:
        # Metadata rows differ between exports (tab or comma). Split heuristically.
        if "\t" in line:
            parts = [p.strip() for p in line.split("\t")]
        else:
            parts = [p.strip() for p in line.split(",")]
        if not parts:
            continue
        key = parts[0].strip().strip(":")
        if key in {"Name", "Owner", "Comments", "Total", "Acquisition", "Acquisition Date"}:
            value = " ".join(p for p in parts[1:] if p)
            meta[key] = value
    return meta


def _find_channel_def_block(lines: List[str], delim: str) -> Tuple[int, int]:
    """Return (start_idx, end_idx_exclusive) for channel definition rows."""
    start = None
    header_tab = "Channel\tName\tFunction"
    header_comma = "Channel,Name,Function"
    for i, line in enumerate(lines):
        s = line.strip()
        if (delim == "\t" and s.startswith(header_tab)) or (delim == "," and s.startswith(header_comma)):
            start = i + 1
            break
    if start is None:
        raise ValueError("Could not find channel definition table header (Channel Name Function...).")

    end = None
    for i in range(start, len(lines)):
        s = lines[i].lstrip()
        # Both formats have a 'Scan Control:' marker line between channel table and data header.
        if s.startswith("Scan") and "Control:" in s:
            end = i
            break
    if end is None:
        raise ValueError("Could not find end of channel definition block (Scan Control:...).")
    return start, end


def _parse_channels(lines: List[str], start: int, end: int, delim: str) -> List[int]:
    channels: List[int] = []
    for line in lines[start:end]:
        parts = _split_fields(line, delim)
        if not parts:
            continue
        try:
            ch = int(str(parts[0]).strip())
        except Exception:
            continue
        channels.append(ch)
    if not channels:
        raise ValueError("No channels found in channel definition block.")
    return channels


def _find_data_header(lines: List[str], delim: str) -> int:
    # The data header typically starts with Scan + Time
    for i, line in enumerate(lines):
        s = line.strip()
        if delim == "\t":
            if s.startswith("Scan\tTime\t"):
                return i
        else:
            if s.startswith("Scan,Time,") or s.startswith("Scan,Time"):
                return i
    raise ValueError("Could not find data table header (Scan Time ...).")


def _parse_timestamp(date_s: str, time_s: str) -> dt.datetime:
    # date: dd/mm/yyyy
    d = dt.datetime.strptime(date_s.strip(), "%d/%m/%Y").date()
    # time: hh:mm:ss:ms
    # Some files may have hh:mm:ss or hh:mm:ss:fff
    m = re.match(r"^(\d{1,2}):(\d{2}):(\d{2})(?::(\d{1,3}))?$", time_s.strip())
    if not m:
        raise ValueError(f"Unrecognised time format: {time_s!r}")
    hh, mm, ss, ms = m.groups()
    micro = int(ms) * 1000 if ms else 0
    t = dt.time(int(hh), int(mm), int(ss), micro)
    return dt.datetime.combine(d, t)


def _parse_timestamp_one_field(dt_s: str) -> dt.datetime:
    # Format examples:
    #  - "19/11/2025 14:30:21:759"
    #  - "19/11/2025 14:30:21"
    s = dt_s.strip()
    if " " not in s:
        raise ValueError(f"Unrecognised datetime format: {dt_s!r}")
    date_s, time_s = s.split(" ", 1)
    return _parse_timestamp(date_s, time_s)


def parse_logger_csv(path: Path) -> ParsedLoggerFile:
    lines = _read_lines_utf16(path)
    delim = _detect_delimiter_from_channel_header(lines)

    meta = _parse_metadata(lines)

    ch_start, ch_end = _find_channel_def_block(lines, delim)
    channels = _parse_channels(lines, ch_start, ch_end, delim)

    data_header_idx = _find_data_header(lines, delim)
    first_data_idx = data_header_idx + 1

    rows: List[Tuple[int, dt.date, dt.time, float, Dict[int, float]]] = []
    first_ts: dt.datetime | None = None

    for line in lines[first_data_idx:]:
        if not line.strip():
            continue

        parts = _split_fields(line, delim)
        if len(parts) < 3:
            continue

        # Stop if we hit another section header
        if parts[0].lower().startswith("scan") and "time" in (parts[1].lower() if len(parts) > 1 else ""):
            continue

        try:
            scan = int(parts[0])
        except Exception:
            continue

        # Two known layouts:
        #  A) Tab export: Scan, Date, Time, (value, alarm)*N
        #  B) Comma export: Scan, DateTime, (value, alarm)*N
        if delim == "\t":
            if len(parts) < 4:
                continue
            date_s = parts[1]
            time_s = parts[2]
            ts = _parse_timestamp(date_s, time_s)
            data_start_idx = 3
        else:
            ts = _parse_timestamp_one_field(parts[1])
            data_start_idx = 2

        if first_ts is None:
            first_ts = ts
        elapsed_min = (ts - first_ts).total_seconds() / 60.0

        values: Dict[int, float] = {}

        # Each channel contributes two columns: value, alarm.
        # Some files include trailing empty columns; ignore.
        for ci, ch in enumerate(channels):
            v_idx = data_start_idx + ci * 2
            if v_idx >= len(parts):
                break
            v = parts[v_idx]
            if v == "":
                continue
            try:
                values[ch] = float(v)
            except Exception:
                continue

        rows.append((scan, ts.date(), ts.time(), elapsed_min, values))

    if not rows:
        raise ValueError("No data rows parsed from file.")

    return ParsedLoggerFile(metadata=meta, channels=channels, rows=rows)

def downsample_full_minutes(parsed: ParsedLoggerFile, tol_seconds: float = 0.6) -> ParsedLoggerFile:
    # Keep row 0 (ambient) and any row where elapsed time is within tol of a whole minute.
    kept: List[Tuple[int, dt.date, dt.time, float, Dict[int, float]]] = []
    for idx, r in enumerate(parsed.rows):
        scan, d, t, elapsed_min, values = r
        if idx == 0:
            kept.append(r)
            continue
        elapsed_sec = elapsed_min * 60.0
        # distance to nearest whole minute
        dist = abs(elapsed_sec - round(elapsed_sec / 60.0) * 60.0)
        if dist <= tol_seconds:
            kept.append((scan, d, t, round(elapsed_sec / 60.0), values))
    return ParsedLoggerFile(metadata=parsed.metadata, channels=parsed.channels, rows=kept)


def _style_header(ws, cell_range: str) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill
            cell.font = font
            cell.alignment = align


def build_report(parsed: ParsedLoggerFile, out_path: Path, source_filename: str) -> None:
    # Categorize channels
    furnace = [ch for ch in parsed.channels if 300 <= ch < 400]
    specimen = [ch for ch in parsed.channels if ch not in furnace]
    furnace.sort()
    specimen.sort()

    # Hard limits per requirement
    specimen = specimen[:35]
    furnace = furnace[:5]

    wb: Workbook = openpyxl.Workbook()
    # Ensure predictable sheet order
    ws_summary = wb.active
    ws_summary.title = "Summary of Results"
    ws_obs = wb.create_sheet("Observations")
    ws_raw = wb.create_sheet("Raw Data")
    ws_cfg = wb.create_sheet("Config")

    # ----------------
    # Config sheet
    # ----------------
    ws_cfg["A1"] = "Group configuration (edit these if your TC layout differs)"
    ws_cfg["A1"].font = Font(bold=True)
    ws_cfg["A3"] = "Face start TC #"
    ws_cfg["A4"] = "Face count"
    ws_cfg["A6"] = "Core start TC #"
    ws_cfg["A7"] = "Core count"
    # Defaults matching the user's example (1-5 face, 6-10 core)
    ws_cfg["B3"] = 1
    ws_cfg["B4"] = min(5, len(specimen))
    ws_cfg["B6"] = 6
    ws_cfg["B7"] = min(5, max(0, len(specimen) - 5))
    ws_cfg.column_dimensions["A"].width = 28
    ws_cfg.column_dimensions["B"].width = 16

    # ----------------
    # Raw Data sheet
    # ----------------
    ws_raw["A1"] = "Imported logger data (absolute and temperature rise)"
    ws_raw["A1"].font = Font(bold=True, size=14)
    # Metadata block
    meta_rows = [
        ("Source file", source_filename),
        ("Name", parsed.metadata.get("Name", "")),
        ("Owner", parsed.metadata.get("Owner", "")),
        ("Acquisition", parsed.metadata.get("Acquisition", "")),
        ("Total channels", str(len(parsed.channels))),
        ("Specimen TCs", str(len(specimen))),
        ("Furnace TCs", str(len(furnace))),
    ]
    start_meta = 3
    for i, (k, v) in enumerate(meta_rows):
        ws_raw.cell(start_meta + i, 1, k).font = Font(bold=True)
        ws_raw.cell(start_meta + i, 2, v)
    ws_raw.column_dimensions["A"].width = 18
    ws_raw.column_dimensions["B"].width = 80

    header_row_1 = 12  # group names
    header_row_2 = 13  # column headers
    data_start_row = 14

    # Base columns
    base_headers = ["Scan", "Date", "Time", "Elapsed (min)"]
    for c, h in enumerate(base_headers, start=1):
        ws_raw.cell(header_row_2, c, h)
        ws_raw.cell(header_row_2, c).alignment = Alignment(horizontal="center", vertical="center")

    # Specimen absolute
    col = 5
    spec_abs_start_col = col
    for i, ch in enumerate(specimen, start=1):
        ws_raw.cell(header_row_1, col, f"TC{i}")
        ws_raw.cell(header_row_2, col, ch)
        col += 1
    spec_abs_end_col = col - 1

    # Specimen rise
    spec_rise_start_col = col
    for i in range(1, len(specimen) + 1):
        ws_raw.cell(header_row_1, col, f"TC{i} ΔT")
        ws_raw.cell(header_row_2, col, f"ΔT{i}")
        col += 1
    spec_rise_end_col = col - 1

    # Furnace absolute
    furnace_abs_start_col = col
    for i, ch in enumerate(furnace, start=1):
        ws_raw.cell(header_row_1, col, f"Furnace TC{i}")
        ws_raw.cell(header_row_2, col, ch)
        col += 1
    furnace_abs_end_col = col - 1

    # Furnace rise
    furnace_rise_start_col = col
    for i in range(1, len(furnace) + 1):
        ws_raw.cell(header_row_1, col, f"Furnace TC{i} ΔT")
        ws_raw.cell(header_row_2, col, f"FΔT{i}")
        col += 1
    furnace_rise_end_col = col - 1

    # Summary columns
    summary_cols = [
        "Mean face ΔT",
        "Max face ΔT",
        "Mean core ΔT",
        "Furnace mean (abs)",
        "Furnace mean ΔT",
    ]
    summary_start_col = col
    for h in summary_cols:
        ws_raw.cell(header_row_1, col, "Summary")
        ws_raw.cell(header_row_2, col, h)
        col += 1
    summary_end_col = col - 1

    # Style headers
    _style_header(ws_raw, f"A{header_row_2}:{get_column_letter(summary_end_col)}{header_row_2}")
    for c in range(1, summary_end_col + 1):
        ws_raw.cell(header_row_1, c).alignment = Alignment(horizontal="center", vertical="center")
    ws_raw.row_dimensions[header_row_1].height = 22
    ws_raw.row_dimensions[header_row_2].height = 28
    ws_raw.freeze_panes = ws_raw["E14"]

    # Column widths
    ws_raw.column_dimensions["C"].width = 12
    ws_raw.column_dimensions["D"].width = 13
    # Hide delta columns to match legacy report layout (keep formulas working)
    for c in range(spec_rise_start_col, spec_rise_end_col + 1):
        ws_raw.column_dimensions[get_column_letter(c)].hidden = True

    for c in range(furnace_rise_start_col, furnace_rise_end_col + 1):
        ws_raw.column_dimensions[get_column_letter(c)].hidden = True


    # Fill data
    for i, (scan, d, t, elapsed_min, values) in enumerate(parsed.rows):
        r = data_start_row + i
        ws_raw.cell(r, 1, scan)
        ws_raw.cell(r, 2, d)
        ws_raw.cell(r, 3, t)
        ws_raw.cell(r, 4, float(elapsed_min))
        ws_raw.cell(r, 2).number_format = "dd/mm/yyyy"
        ws_raw.cell(r, 3).number_format = "hh:mm:ss"
        ws_raw.cell(r, 4).number_format = "0"

        # specimen absolute
        for j, ch in enumerate(specimen):
            v = values.get(ch)
            ws_raw.cell(r, spec_abs_start_col + j, v)

        # furnace absolute
        for j, ch in enumerate(furnace):
            v = values.get(ch)
            ws_raw.cell(r, furnace_abs_start_col + j, v)

    last_data_row = data_start_row + len(parsed.rows) - 1

    # Formulas for rises and summaries
    ambient_row = data_start_row  # first data row is ambient
    for r in range(data_start_row, last_data_row + 1):
        # specimen rises
        for j in range(len(specimen)):
            abs_col = spec_abs_start_col + j
            rise_col = spec_rise_start_col + j
            abs_addr = f"{get_column_letter(abs_col)}{r}"
            amb_addr = f"{get_column_letter(abs_col)}{ambient_row}"
            ws_raw.cell(r, rise_col).value = f"=IF({abs_addr}=\"\",\"\",{abs_addr}-{amb_addr})"
            ws_raw.cell(r, rise_col).number_format = "0.0"

        # furnace rises
        for j in range(len(furnace)):
            abs_col = furnace_abs_start_col + j
            rise_col = furnace_rise_start_col + j
            abs_addr = f"{get_column_letter(abs_col)}{r}"
            amb_addr = f"{get_column_letter(abs_col)}{ambient_row}"
            ws_raw.cell(r, rise_col).value = f"=IF({abs_addr}=\"\",\"\",{abs_addr}-{amb_addr})"
            ws_raw.cell(r, rise_col).number_format = "0.0"

        # mean/max/core using OFFSET over specimen rise columns
        tc1_rise_addr = f"{get_column_letter(spec_rise_start_col)}{r}"
        mean_face_col = summary_start_col
        max_face_col = summary_start_col + 1
        mean_core_col = summary_start_col + 2
        furnace_mean_abs_col = summary_start_col + 3
        furnace_mean_rise_col = summary_start_col + 4

        ws_raw.cell(r, mean_face_col).value = (
            f"=IF(Config!$B$4=0,\"\",AVERAGE(OFFSET({tc1_rise_addr},0,Config!$B$3-1,1,Config!$B$4)))"
        )
        ws_raw.cell(r, max_face_col).value = (
            f"=IF(Config!$B$4=0,\"\",MAX(OFFSET({tc1_rise_addr},0,Config!$B$3-1,1,Config!$B$4)))"
        )
        ws_raw.cell(r, mean_core_col).value = (
            f"=IF(Config!$B$7=0,\"\",AVERAGE(OFFSET({tc1_rise_addr},0,Config!$B$6-1,1,Config!$B$7)))"
        )

        # Furnace mean (abs) + rise
        if len(furnace) > 0:
            abs_start = f"{get_column_letter(furnace_abs_start_col)}{r}"
            abs_end = f"{get_column_letter(furnace_abs_end_col)}{r}"
            rise_start = f"{get_column_letter(furnace_rise_start_col)}{r}"
            rise_end = f"{get_column_letter(furnace_rise_end_col)}{r}"
            ws_raw.cell(r, furnace_mean_abs_col).value = f"=AVERAGE({abs_start}:{abs_end})"
            ws_raw.cell(r, furnace_mean_rise_col).value = f"=AVERAGE({rise_start}:{rise_end})"
        else:
            ws_raw.cell(r, furnace_mean_abs_col).value = ""
            ws_raw.cell(r, furnace_mean_rise_col).value = ""

        for c in range(summary_start_col, summary_end_col + 1):
            ws_raw.cell(r, c).number_format = "0.0"

    # ----------------
    # Summary sheet
    # ----------------
    ws_summary["A1"] = "Test summary (auto-generated)"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Source file"
    ws_summary["B3"] = source_filename
    ws_summary["A4"] = "Total specimen TCs"
    ws_summary["B4"] = len(specimen)
    ws_summary["A5"] = "Total furnace TCs"
    ws_summary["B5"] = len(furnace)
    ws_summary["A6"] = "Note"
    ws_summary["B6"] = "Edit Config tab if face/core grouping differs (defaults: face 1-5, core 6-10)."
    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 70

    # Create charts
    def add_line_chart(title: str, y_col: int, anchor: str) -> None:
        chart = LineChart()
        chart.title = title
        chart.style = 2
        chart.y_axis.title = "Temperature rise (°C)"
        chart.x_axis.title = "Elapsed (min)"

        xref = Reference(ws_raw, min_col=4, min_row=data_start_row, max_row=last_data_row)
        yref = Reference(ws_raw, min_col=y_col, min_row=data_start_row, max_row=last_data_row)
        chart.add_data(yref, titles_from_data=False)
        chart.set_categories(xref)
        chart.legend = None
        chart.height = 8
        chart.width = 18
        ws_summary.add_chart(chart, anchor)

    add_line_chart("Mean temperature rise (face)", summary_start_col, "A9")
    add_line_chart("Maximum temperature rise (face)", summary_start_col + 1, "A25")
    add_line_chart("Mean temperature rise (core)", summary_start_col + 2, "A41")

    # ----------------
    # Observations sheet
    # ----------------
    ws_obs["A1"] = "Observations"
    ws_obs["A1"].font = Font(bold=True, size=14)
    ws_obs["A3"] = "(This tab is intentionally free-form. Paste or type your test-specific notes here.)"
    ws_obs["A3"].alignment = Alignment(wrap_text=True)
    ws_obs.column_dimensions["A"].width = 100

    # Slightly nicer default view
    for ws in [ws_summary, ws_obs, ws_raw, ws_cfg]:
        ws.sheet_view.showGridLines = False

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert logger CSV to Excel report")
    parser.add_argument("input_csv", type=str)
    parser.add_argument("output_xlsx", type=str)
    args = parser.parse_args()

    inp = Path(args.input_csv)
    out = Path(args.output_xlsx)

    parsed = parse_logger_csv(inp)
    parsed_ds = downsample_full_minutes(parsed)
    build_report(parsed_ds, out, source_filename=inp.name)


if __name__ == "__main__":
    main()
