# app.py

import os
import traceback
import tkinter as tk
import openpyxl
from tkinter import filedialog, messagebox
from pathlib import Path
from logger_to_report import parse_logger_csv, downsample_full_minutes, build_report

# If you already have a generator module, import it here.
# Example (if you have logger_to_report.py with a function you can call):
# from logger_to_report import generate_report


def build_output_path(input_csv_path: str) -> str:
    folder = os.path.dirname(input_csv_path)
    base = os.path.splitext(os.path.basename(input_csv_path))[0]
    return os.path.join(folder, f"{base} - report.xlsx")


def generate_report(
    input_csv_path: str,
    output_xlsx_path: str,
    face_start: int,
    face_count: int,
    core_start: int,
    core_count: int,
    furnace_min: int,
    furnace_max: int,
    minute_tolerance_seconds: float,
) -> None:
    # Read + parse
    parsed = parse_logger_csv(Path(input_csv_path))

    # Downsample to whole minutes
    parsed = downsample_full_minutes(parsed, tol_seconds=minute_tolerance_seconds)

    # Build the report workbook (this writes the file)
    build_report(parsed, Path(output_xlsx_path), source_filename=os.path.basename(input_csv_path))

    # Apply UI settings into the Config sheet after creation
    wb = openpyxl.load_workbook(output_xlsx_path)
    if "Config" in wb.sheetnames:
        ws = wb["Config"]
        ws["B3"] = face_start
        ws["B4"] = face_count
        ws["B6"] = core_start
        ws["B7"] = core_count
    wb.save(output_xlsx_path)


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("INSTR Logger → Excel Report")
        self.resizable(False, False)

        self.input_path_var = tk.StringVar(value="")
        self.output_path_var = tk.StringVar(value="")

        # Defaults based on your typical setup
        self.face_start_var = tk.StringVar(value="1")
        self.face_count_var = tk.StringVar(value="5")
        self.core_start_var = tk.StringVar(value="6")
        self.core_count_var = tk.StringVar(value="5")
        self.furnace_min_var = tk.StringVar(value="300")
        self.furnace_max_var = tk.StringVar(value="399")
        self.minute_tol_var = tk.StringVar(value="0.5")

        self._build_ui()

    def _build_ui(self) -> None:
        pad = 8

        # Input row
        frm_file = tk.Frame(self)
        frm_file.pack(fill="x", padx=pad, pady=(pad, 0))

        tk.Label(frm_file, text="Input CSV:").pack(anchor="w")
        row = tk.Frame(frm_file)
        row.pack(fill="x")

        entry_in = tk.Entry(row, textvariable=self.input_path_var, width=72)
        entry_in.pack(side="left", expand=True, fill="x")

        tk.Button(row, text="Browse...", command=self.browse_input).pack(side="left", padx=(6, 0))

        # Output row (auto)
        frm_out = tk.Frame(self)
        frm_out.pack(fill="x", padx=pad, pady=(pad, 0))

        tk.Label(frm_out, text="Output (auto):").pack(anchor="w")
        tk.Entry(frm_out, textvariable=self.output_path_var, width=72, state="readonly").pack(fill="x")

        # Options
        frm_opts = tk.LabelFrame(self, text="Options")
        frm_opts.pack(fill="x", padx=pad, pady=pad)

        grid = tk.Frame(frm_opts)
        grid.pack(fill="x", padx=pad, pady=pad)

        def add_row(label: str, var: tk.StringVar, r: int) -> None:
            tk.Label(grid, text=label, width=22, anchor="w").grid(row=r, column=0, sticky="w", pady=2)
            tk.Entry(grid, textvariable=var, width=12).grid(row=r, column=1, sticky="w", pady=2)

        add_row("Face TC start:", self.face_start_var, 0)
        add_row("Face TC count:", self.face_count_var, 1)
        add_row("Core TC start:", self.core_start_var, 2)
        add_row("Core TC count:", self.core_count_var, 3)
        add_row("Furnace min channel:", self.furnace_min_var, 4)
        add_row("Furnace max channel:", self.furnace_max_var, 5)
        add_row("Minute tolerance (sec):", self.minute_tol_var, 6)

        # Actions
        frm_actions = tk.Frame(self)
        frm_actions.pack(fill="x", padx=pad, pady=(0, pad))

        self.run_btn = tk.Button(frm_actions, text="Generate Report", command=self.on_run, state="disabled")
        self.run_btn.pack(side="right")

        tk.Button(frm_actions, text="Exit", command=self.destroy).pack(side="right", padx=(0, 8))

    def browse_input(self) -> None:
        path = filedialog.askopenfilename(
            title="Select INSTR CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return

        self.input_path_var.set(path)
        self.output_path_var.set(build_output_path(path))
        self.run_btn.config(state="normal")

    def _get_int(self, label: str, value: str) -> int:
        try:
            return int(value)
        except ValueError:
            raise ValueError(f"{label} must be an integer.")

    def _get_float(self, label: str, value: str) -> float:
        try:
            return float(value)
        except ValueError:
            raise ValueError(f"{label} must be a number.")

    def on_run(self) -> None:
        input_path = self.input_path_var.get().strip()
        output_path = self.output_path_var.get().strip()

        if not input_path:
            messagebox.showerror("Missing input", "Please select an input CSV file.")
            return

        try:
            face_start = self._get_int("Face TC start", self.face_start_var.get().strip())
            face_count = self._get_int("Face TC count", self.face_count_var.get().strip())
            core_start = self._get_int("Core TC start", self.core_start_var.get().strip())
            core_count = self._get_int("Core TC count", self.core_count_var.get().strip())
            furnace_min = self._get_int("Furnace min channel", self.furnace_min_var.get().strip())
            furnace_max = self._get_int("Furnace max channel", self.furnace_max_var.get().strip())
            minute_tol = self._get_float("Minute tolerance (sec)", self.minute_tol_var.get().strip())

            self.run_btn.config(state="disabled")
            self.update_idletasks()

            generate_report(
                input_csv_path=input_path,
                output_xlsx_path=output_path,
                face_start=face_start,
                face_count=face_count,
                core_start=core_start,
                core_count=core_count,
                furnace_min=furnace_min,
                furnace_max=furnace_max,
                minute_tolerance_seconds=minute_tol,
            )

            messagebox.showinfo("Done", f"Report created:\n{output_path}")

        except Exception as ex:
            details = traceback.format_exc()
            messagebox.showerror("Error", f"{ex}\n\nDetails:\n{details}")
        finally:
            self.run_btn.config(state="normal")


if __name__ == "__main__":
    app = App()
    app.mainloop()
